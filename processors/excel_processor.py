"""
Excel处理器
处理Excel文件，支持.xls和.xlsx格式，包含合并单元格处理
"""

from pathlib import Path
from typing import Dict, Any

from .base import BaseProcessor, ProcessingResult

# Excel处理相关导入
try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    import xlrd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelProcessor(BaseProcessor):
    """Excel文档处理器"""
    
    def __init__(self, config: Dict[str, Any] = None):
        """
        初始化Excel处理器
        
        Args:
            config: 配置字典
        """
        super().__init__(config)
        
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel处理相关库未安装，无法处理Excel文件")
    
    def get_supported_extensions(self) -> list:
        """获取支持的文件扩展名"""
        return ['.xls', '.xlsx']
    
    def parse_xlsx_file(self, file_path: Path, active_sheet_only: bool = False) -> Dict[str, Any]:
        """
        使用openpyxl处理.xlsx文件
        
        Args:
            file_path: Excel文件路径
            active_sheet_only: 是否只处理活动工作表
            
        Returns:
            Dict: {工作表名: DataFrame}
        """
        wb = load_workbook(file_path)
        sheets_data = {}
        
        # 确定要处理的工作表列表
        sheets_to_process = [wb.active] if active_sheet_only else wb.worksheets
        
        for ws in sheets_to_process:
            # 创建合并单元格映射
            merged_map = {}
            if ws.merged_cells.ranges:
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            merged_map[(row, col)] = (min_row, min_col)
            
            # 构建完整数据
            data = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    if (cell.row, cell.column) in merged_map:
                        m_row, m_col = merged_map[(cell.row, cell.column)]
                        if cell.row == m_row and cell.column == m_col:
                            row_data.append(cell.value)
                        else:
                            row_data.append("")  # 标记为合并单元格
                    else:
                        row_data.append(cell.value)
                data.append(row_data)
            
            # 创建DataFrame
            if data:
                # 处理列名
                headers = data[0] if data else []
                # 确保列名不为None
                headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(headers)]
                
                # 创建DataFrame
                df_data = data[1:] if len(data) > 1 else []
                df = pd.DataFrame(df_data, columns=headers)
                sheets_data[ws.title] = df
        
        return sheets_data
    
    def parse_xls_file(self, file_path: Path, active_sheet_only: bool = False) -> Dict[str, Any]:
        """
        使用xlrd处理.xls文件
        
        Args:
            file_path: Excel文件路径
            active_sheet_only: 是否只处理活动工作表
            
        Returns:
            Dict: {工作表名: DataFrame}
        """
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        sheets_data = {}
        
        # 确定要处理的工作表列表
        if active_sheet_only:
            sheets_to_process = [wb.sheet_by_index(wb.active_sheet)]
        else:
            sheets_to_process = [wb.sheet_by_index(i) for i in range(wb.nsheets)]
        
        for ws in sheets_to_process:
            # 获取合并单元格信息
            merged_cells = ws.merged_cells
            merged_map = {}
            
            for crange in merged_cells:
                rlo, rhi, clo, chi = crange
                for row in range(rlo, rhi):
                    for col in range(clo, chi):
                        merged_map[(row, col)] = (rlo, clo)
            
            # 构建完整数据
            data = []
            for row_idx in range(ws.nrows):
                row_data = []
                for col_idx in range(ws.ncols):
                    if (row_idx, col_idx) in merged_map:
                        m_row, m_col = merged_map[(row_idx, col_idx)]
                        if row_idx == m_row and col_idx == m_col:
                            row_data.append(ws.cell_value(row_idx, col_idx))
                        else:
                            row_data.append("")  # 标记为合并单元格
                    else:
                        row_data.append(ws.cell_value(row_idx, col_idx))
                data.append(row_data)
            
            # 创建DataFrame
            if data:
                # 处理列名
                headers = data[0] if data else []
                headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(headers)]
                
                # 创建DataFrame
                df_data = data[1:] if len(data) > 1 else []
                df = pd.DataFrame(df_data, columns=headers)
                sheets_data[ws.name] = df
        
        return sheets_data
    
    def parse_excel_with_merged_cells(self, file_path: Path, active_sheet_only: bool = False) -> Dict[str, Any]:
        """
        解析包含合并单元格的Excel文件
        
        Args:
            file_path: Excel文件路径
            active_sheet_only: 是否只处理活动工作表
            
        Returns:
            Dict: {工作表名: DataFrame}
        """
        file_ext = file_path.suffix.lower()
        
        if file_ext == '.xls':
            return self.parse_xls_file(file_path, active_sheet_only)
        else:
            return self.parse_xlsx_file(file_path, active_sheet_only)
    
    def convert_to_markdown(self, sheets_data: Dict[str, Any], file_name: str) -> str:
        """
        将Excel数据转换为Markdown格式
        
        Args:
            sheets_data: 工作表数据字典
            file_name: 文件名
            
        Returns:
            str: Markdown格式的内容
        """
        markdown_content = f"# {file_name}\n\n"
        
        for sheet_name, df in sheets_data.items():
            markdown_content += f"## {sheet_name}\n\n"
            
            if not df.empty:
                # 转换为Markdown表格
                try:
                    markdown_table = df.to_markdown(index=False)
                    markdown_content += markdown_table
                except Exception as e:
                    self.logger.warning(f"转换工作表 {sheet_name} 为Markdown时出错: {e}")
                    # 回退到简单的文本格式
                    markdown_content += f"```\n{df.to_string(index=False)}\n```"
            else:
                markdown_content += "*此工作表为空*"
            
            markdown_content += "\n\n"
        
        return markdown_content.strip()
    
    def process(self, file_path: Path) -> ProcessingResult:
        """
        处理Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            ProcessingResult: 处理结果
        """
        try:
            self.logger.info(f"开始处理Excel文件: {file_path.name}")
            
            # 解析Excel文件
            sheets_data = self.parse_excel_with_merged_cells(file_path)
            
            if not sheets_data:
                return ProcessingResult(
                    success=False,
                    error="Excel文件中没有找到有效的工作表"
                )
            
            # 转换为Markdown
            file_stem = file_path.stem
            markdown_content = self.convert_to_markdown(sheets_data, file_stem)
            
            # 构建元数据
            metadata = {
                'file_type': 'excel',
                'sheets_count': len(sheets_data),
                'sheet_names': list(sheets_data.keys()),
                'file_extension': file_path.suffix.lower()
            }
            
            return ProcessingResult(
                success=True,
                content=markdown_content,
                metadata=metadata
            )
            
        except Exception as e:
            error_msg = f"处理Excel文件失败: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            return ProcessingResult(
                success=False,
                error=error_msg
            )
